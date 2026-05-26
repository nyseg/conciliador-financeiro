import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Cores
COR_OK = "E1F5EE"
COR_ERR = "FCEBEB"
COR_WARN = "FAEEDA"
COR_HEADER = "1A1A2E"
COR_SUBHEADER = "E8E8F0"

def gerar_excel(payload: dict) -> bytes:
    wb = Workbook()
    modo = payload.get('modo', 'despesas')

    # Aba Resumo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    _preencher_resumo(ws_resumo, payload)

    # Aba Itens
    ws_itens = wb.create_sheet("Conciliação Detalhada")
    if modo == 'despesas':
        _preencher_itens_despesas(ws_itens, payload.get('itens', []))
    else:
        _preencher_itens_receitas(ws_itens, payload.get('itens', []))

    # Aba Encargos (apenas despesas)
    if modo == 'despesas' and payload.get('encargos_pendentes'):
        ws_enc = wb.create_sheet("Encargos Pendentes")
        _preencher_encargos(ws_enc, payload['encargos_pendentes'])

    # Aba Faturas ERP (apenas despesas)
    if modo == 'despesas' and payload.get('faturas_erp'):
        ws_fat = wb.create_sheet("Faturas ERP")
        _preencher_faturas_erp(ws_fat, payload['faturas_erp'])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _preencher_resumo(ws, payload):
    resumo = payload.get('resumo', {})
    modo = payload.get('modo', 'despesas')
    periodo = payload.get('periodo', '')

    _header_cell(ws, "A1", "CONCILIADOR FINANCEIRO", size=14)
    ws.merge_cells("A1:D1")
    _header_cell(ws, "A2", f"Período: {periodo or 'Todos'}  |  Modo: {modo.upper()}", size=11, bg=COR_SUBHEADER, color="333333")
    ws.merge_cells("A2:D2")

    ws["A4"] = "Indicador"
    ws["B4"] = "Valor"
    for cell in [ws["A4"], ws["B4"]]:
        _estilo_header(cell)

    linhas = []
    if modo == 'despesas':
        linhas = [
            ("Total de itens analisados", resumo.get('total_itens', 0)),
            ("Itens conciliados", resumo.get('conciliados', 0)),
            ("Na fatura, sem ERP", resumo.get('sem_erp', 0)),
            ("No ERP, sem fatura", resumo.get('sem_fatura', 0)),
            ("Total fatura cartão (R$)", f"R$ {resumo.get('total_fatura', 0):.2f}"),
            ("Total lançado ERP (R$)", f"R$ {resumo.get('total_erp', 0):.2f}"),
            ("Diferença (R$)", f"R$ {resumo.get('diferenca', 0):.2f}"),
            ("Encargos não lançados (R$)", f"R$ {resumo.get('total_encargos_pendentes', 0):.2f}"),
        ]
    else:
        linhas = [
            ("Total de itens analisados", resumo.get('total_itens', 0)),
            ("Itens conciliados", resumo.get('conciliados', 0)),
            ("Divergências", resumo.get('divergencias', 0)),
            ("Ausentes", resumo.get('ausentes', 0)),
            ("Total operadora (R$)", f"R$ {resumo.get('total_operadora', 0):.2f}"),
            ("Total ERP (R$)", f"R$ {resumo.get('total_erp', 0):.2f}"),
            ("Total banco (R$)", f"R$ {resumo.get('total_banco', 0):.2f}"),
            ("Diferença op. vs ERP (R$)", f"R$ {resumo.get('diferenca_op_erp', 0):.2f}"),
        ]

    for i, (label, valor) in enumerate(linhas, start=5):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = valor
        ws[f"A{i}"].alignment = Alignment(horizontal="left")
        ws[f"B{i}"].alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20


def _preencher_itens_despesas(ws, itens):
    headers = ["Data Fatura", "Estabelecimento / Encargo", "Cartão", "Valor Fatura (R$)",
               "Data ERP", "Categoria ERP", "Valor ERP (R$)", "Nº Fatura ERP", "Sit. ERP", "Status"]
    _escrever_headers(ws, headers)

    for i, item in enumerate(itens, start=2):
        status = item.get('status', '')
        cor = COR_OK if status == 'ok' else COR_ERR if status == 'ausente_erp' else COR_WARN

        vals = [
            item.get('data_fatura', '—'),
            item.get('descricao_fatura', '—'),
            item.get('cartao', '—'),
            item.get('valor_fatura', 0),
            item.get('data_erp', '—'),
            item.get('descricao_erp', '—'),
            item.get('valor_erp', 0),
            item.get('numero_fatura_erp', '—'),
            item.get('status_erp', '—'),
            _status_label(status),
        ]
        for j, val in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.fill = PatternFill(fill_type="solid", fgColor=cor)
            if j in [4, 7]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    _ajustar_colunas(ws, [18, 30, 10, 16, 18, 25, 16, 15, 12, 15])


def _preencher_itens_receitas(ws, itens):
    headers = ["Data Operadora", "Descrição", "Valor Operadora (R$)",
               "Data ERP", "Descrição ERP", "Valor ERP (R$)", "Valor Banco (R$)", "Status"]
    _escrever_headers(ws, headers)

    for i, item in enumerate(itens, start=2):
        status = item.get('status', '')
        cor = COR_OK if status == 'ok' else COR_WARN if status == 'divergencia' else COR_ERR
        vals = [
            item.get('data_operadora', '—'),
            item.get('descricao_operadora', '—'),
            item.get('valor_operadora', 0),
            item.get('data_erp', '—'),
            item.get('descricao_erp', '—'),
            item.get('valor_erp', 0),
            item.get('valor_banco', 0),
            _status_label(status),
        ]
        for j, val in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.fill = PatternFill(fill_type="solid", fgColor=cor)
            if j in [3, 6, 7]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    _ajustar_colunas(ws, [18, 30, 18, 18, 25, 16, 16, 15])


def _preencher_encargos(ws, encargos):
    headers = ["Encargo", "Cartão", "Valor (R$)", "Lançado no ERP", "Ação Necessária"]
    _escrever_headers(ws, headers)
    for i, enc in enumerate(encargos, start=2):
        vals = [
            enc.get('descricao', ''),
            enc.get('cartao', ''),
            enc.get('valor', 0),
            "Sim" if enc.get('lançado_erp') else "Não",
            "Lançar no ERP" if not enc.get('lançado_erp') else "OK",
        ]
        for j, val in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.fill = PatternFill(fill_type="solid", fgColor=COR_ERR if not enc.get('lançado_erp') else COR_OK)
            if j == 3:
                cell.number_format = '#,##0.00'
    _ajustar_colunas(ws, [28, 10, 16, 18, 20])


def _preencher_faturas_erp(ws, faturas):
    headers = ["Nº Fatura", "Período", "Vencimento", "Total ERP (R$)", "Qtd Itens", "Status"]
    _escrever_headers(ws, headers)
    for i, fat in enumerate(faturas, start=2):
        vals = [
            fat.get('numero_fatura', ''),
            fat.get('periodo', ''),
            fat.get('vencimento', ''),
            fat.get('total_erp', 0),
            fat.get('qtd_itens', 0),
            fat.get('status', ''),
        ]
        cor = COR_WARN if fat.get('status') == 'Pendente' else COR_OK
        for j, val in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.fill = PatternFill(fill_type="solid", fgColor=cor)
    _ajustar_colunas(ws, [14, 20, 16, 16, 12, 12])


# ── HELPERS ──────────────────────────────────────────────────────────────────

def _header_cell(ws, ref, text, size=12, bg=COR_HEADER, color="FFFFFF"):
    cell = ws[ref]
    cell.value = text
    cell.font = Font(bold=True, size=size, color=color)
    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _estilo_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor=COR_HEADER)
    cell.alignment = Alignment(horizontal="center")


def _escrever_headers(ws, headers):
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        _estilo_header(cell)


def _ajustar_colunas(ws, larguras):
    for j, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _status_label(status):
    return {"ok": "✅ Conciliado", "ausente_erp": "❌ Sem ERP",
            "ausente_fatura": "⚠️ Sem fatura", "divergencia": "⚠️ Divergência"}.get(status, status)
