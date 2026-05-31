import pandas as pd
import re
import io
from openpyxl import load_workbook
from typing import Optional
import pdfplumber

# ── OFX / QFX PARSER ────────────────────────────────────────────────────────

def _parsear_ofx(conteudo: bytes) -> pd.DataFrame:
    """
    Parseia arquivo OFX/QFX (formato SGML 1.x ou XML 2.x).
    Extrai transações (STMTTRN) de extratos bancários e faturas de cartão.
    Retorna DataFrame com colunas: data, descricao, valor, nsu, tipo.
    """
    try:
        text = conteudo.decode("utf-8", errors="replace")
    except Exception:
        text = conteudo.decode("latin-1", errors="replace")

    def _get(tag: str, block: str) -> str:
        m = re.search(fr"<{tag}[^>]*>(.*?)(?=<|\Z)", block, re.IGNORECASE | re.DOTALL)
        return m.group(1).strip() if m else ""

    rows = []
    for block in re.findall(r"<STMTTRN>(.*?)</STMTTRN>", text, re.DOTALL | re.IGNORECASE):
        trnamt  = _get("TRNAMT", block)
        dtpost  = _get("DTPOSTED", block)
        memo    = _get("MEMO", block) or _get("NAME", block) or ""
        fitid   = _get("FITID", block)
        trntype = _get("TRNTYPE", block)

        try:
            valor = float(trnamt.replace(",", "."))
        except (ValueError, TypeError):
            continue

        dt = None
        if dtpost:
            try:
                dt = pd.to_datetime(dtpost[:8], format="%Y%m%d")
            except Exception:
                pass

        rows.append({
            "data": dt,
            "descricao": memo.strip(),
            "valor": abs(round(valor, 2)),
            "nsu": fitid,
            "tipo": trntype.lower() if trntype else ("debito" if valor < 0 else "credito"),
        })

    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ── PDF PARSER ───────────────────────────────────────────────────────────────

def _inferir_ano(lines: list) -> Optional[int]:
    """
    Tenta extrair o ano de referência de linhas de texto do PDF.
    Procura por datas completas (DD/MM/AAAA) em qualquer lugar do texto.
    Útil para faturas Santander que usam DD/MM nas transações sem ano.
    """
    import datetime
    re_ano = re.compile(r'\d{2}[\/\-]\d{2}[\/\-](\d{4})')
    for line in lines:
        m = re_ano.search(line)
        if m:
            ano = int(m.group(1))
            if 2020 <= ano <= 2035:
                return ano
    return datetime.date.today().year

# Palavras-chave para ignorar linhas de totais/cabeçalhos no OCR
_OCR_SKIP_KW = {
    'total da fatura', 'total nacional', 'total internacional',
    'subtotal', 'saldo anterior', 'saldo devedor',
    'encargos', 'juros', 'iof', 'multa',
    'limite de crédito', 'crédito disponível',
    'data de vencimento', 'pagamento mínimo',
    'valores em reais', 'lançamentos nacionais', 'lançamentos internacionais',
    'próxima fatura', 'fatura anterior',
}

# ── Regexes para datas ────────────────────────────────────────────────────────
# Formato completo: 15/03/2024 ou 15-03-2024
_RE_DATE_FULL  = re.compile(r'^(\d{2}[\/\-]\d{2}[\/\-]\d{2,4})\s+')
# Formato Santander: 15/03 (sem ano) — valor do ano inferido depois
_RE_DATE_SHORT = re.compile(r'^(\d{2}[\/\-]\d{2})\s+(?!\d{2,4}[\s\/\-])')

# ── Regex para valor monetário no final da linha ──────────────────────────────
# Suporta: 1.500,00 | 1500,00 | 1 500,00 | R$ 1.500,00
_RE_VALUE_END = re.compile(
    r'R?\$?\s*([\d]{1,3}(?:[.\s]\d{3})*[,]\d{2}|[\d]+[,]\d{2})\s*$'
)

# UF brasileiras — Santander inclui cidade/UF no final da descrição
_UF_BR = re.compile(
    r'\s+[A-Z]{2,}\s+(AC|AL|AP|AM|BA|CE|DF|ES|GO|MA|MT|MS|MG|PA|PB|PE|PI|'
    r'RJ|RN|RS|RO|RR|SC|SP|SE|TO|PR)\s*$'
)


def _parse_linhas_transacao(lines: list, ano_referencia: Optional[int] = None) -> list:
    """
    Analisa linhas de texto (pdfplumber ou OCR) e retorna transações.

    Suporta:
      - Datas completas: 15/03/2024
      - Datas Santander: 15/03 (sem ano — usa ano_referencia ou ano atual)
      - Valores BR: 1.500,00 | 1500,00
      - Descrições com cidade/UF no final (removidas)
    """
    import datetime
    ano = ano_referencia or datetime.date.today().year

    rows = []
    for line in lines:
        line = line.strip()
        if not line or len(line) < 6:
            continue

        # Pula cabeçalhos e totais (verificação parcial — contém a keyword)
        ll = line.lower()
        if any(k in ll for k in _OCR_SKIP_KW):
            continue

        # Tenta data completa primeiro, depois data curta (Santander)
        date_m = _RE_DATE_FULL.match(line) or _RE_DATE_SHORT.match(line)
        if not date_m:
            continue

        value_m = _RE_VALUE_END.search(line)
        if not value_m:
            continue

        date_str  = date_m.group(1)
        value_str = value_m.group(1).replace(' ', '')

        # Completa data curta com ano de referência
        if len(date_str) <= 5:          # "15/03"
            date_str = f"{date_str}/{ano}"

        desc_start = date_m.end()
        # Posição do valor no final da linha (sem o "R$ " prefixo)
        value_pos  = line.rfind(value_str)
        if value_pos <= desc_start:
            continue

        descricao = line[desc_start:value_pos].strip()
        if not descricao or len(descricao) < 2:
            continue

        # Remove cidade/UF do final da descrição (ex: "SAO PAULO SP")
        descricao = _UF_BR.sub('', descricao).strip()
        # Remove trailing dígitos isolados que sobram (ex: código de loja "12345")
        descricao = re.sub(r'\s+\d{4,}\s*$', '', descricao).strip()

        rows.append({
            'data':      date_str,
            'descricao': descricao,
            'valor':     value_str,
        })
    return rows


def _extrair_texto_pymupdf(conteudo: bytes) -> list:
    """
    Extrai linhas de texto via PyMuPDF (fitz).
    PyMuPDF usa engine própria, diferente do pdfplumber, e consegue
    decodificar fontes customizadas que o pdfplumber não consegue.
    Retorna lista de linhas de texto.
    """
    try:
        import fitz
    except ImportError:
        return []
    lines = []
    try:
        doc = fitz.open(stream=conteudo, filetype='pdf')
        for page in doc:
            text = page.get_text("text") or ""
            lines.extend(text.splitlines())
    except Exception:
        pass
    return lines


def _parsear_pdf_ocr(conteudo: bytes) -> pd.DataFrame:
    """
    OCR via PyMuPDF + pytesseract: renderiza páginas como imagem e lê com OCR.
    Usado para PDFs totalmente baseados em imagem (scans ou faturas rasterizadas).
    Lança ValueError com mensagem clara se Tesseract não estiver instalado.
    """
    try:
        import fitz
        from PIL import Image as _PILImage
        from PIL import ImageFilter, ImageOps
    except ImportError:
        raise ValueError(
            "PyMuPDF/Pillow não instalado no servidor. Execute o redeploy no Render."
        )

    try:
        import pytesseract
    except ImportError:
        raise ValueError(
            "pytesseract não instalado. Execute o redeploy no Render."
        )

    def _preprocessar(img):
        """Melhora contraste e nitidez para OCR mais preciso."""
        img = ImageOps.autocontrast(img, cutoff=1)
        img = img.filter(ImageFilter.SHARPEN)
        return img

    def _ocr_page(img) -> str:
        """
        Tenta OCR em português+inglês. Lança TesseractNotFoundError se
        o binário do Tesseract não estiver instalado no sistema.
        """
        img = _preprocessar(img)
        # Renderiza em RGB e converte para L (cinza) — melhor qualidade CMYK → cinza
        if img.mode != 'L':
            img = img.convert('L')
        for lang in ('por+eng', 'por', 'eng'):
            try:
                return pytesseract.image_to_string(
                    img, lang=lang,
                    config='--psm 6 --oem 3',
                )
            except pytesseract.TesseractNotFoundError:
                raise   # propaga — não é problema de idioma, é de instalação
            except Exception:
                continue
        return ''

    all_lines: list = []
    doc = fitz.open(stream=conteudo, filetype='pdf')
    for page_num in range(len(doc)):
        page = doc[page_num]
        # Renderiza em RGB (melhor conversão de CMYK) a ~216 DPI
        mat = fitz.Matrix(3, 3)
        pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
        img = _PILImage.frombytes('RGB', (pix.width, pix.height), pix.samples)
        text = _ocr_page(img)
        all_lines.extend(text.splitlines())

    ano_ref = _inferir_ano(all_lines)
    rows = _parse_linhas_transacao(all_lines, ano_referencia=ano_ref)
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _parsear_pdf(conteudo: bytes) -> pd.DataFrame:
    """
    Parseia PDF financeiro em 4 camadas progressivas:
      1. pdfplumber  — tabelas com bordas (PDFs tabulares)
      2. pdfplumber  — texto linha a linha (PDFs semi-estruturados)
      3. PyMuPDF     — engine alternativa, decodifica fontes customizadas
      4. OCR         — PyMuPDF + Tesseract (PDFs baseados em imagem/scan)

    Lança ValueError se nenhum método extraiu transações.
    """
    # ── Camada 1: tabelas estruturadas (pdfplumber) ───────────────────────────
    all_rows: list = []
    headers:  list | None = None

    with pdfplumber.open(io.BytesIO(conteudo)) as pdf:
        for page in pdf.pages:
            for tset in [
                {"vertical_strategy": "lines_strict", "horizontal_strategy": "lines_strict", "snap_tolerance": 4},
                {},   # estratégia padrão permissiva
            ]:
                tables = page.extract_tables(table_settings=tset) if tset else page.extract_tables()
                tables = tables or []
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    raw_header = [str(c or "").strip() for c in table[0]]
                    # Rejeita tabelas onde o header parece ser um resumo financeiro
                    # (ex: "Fatura Anterior", "Total", "Saldo") e não colunas de dados
                    palavras_resumo = {'fatura anterior', 'total', 'saldo', 'pagamento',
                                       'limite', 'encargo', 'juros', 'vencimento'}
                    if all(h.lower() in palavras_resumo or h == '' for h in raw_header):
                        continue
                    if headers is None:
                        headers   = [h or f"col_{i}" for i, h in enumerate(raw_header)]
                        data_rows = table[1:]
                    else:
                        norm_raw    = [h.lower() for h in raw_header]
                        norm_header = [h.lower() for h in headers]
                        data_rows   = table[1:] if norm_raw == norm_header else table
                    for row in data_rows:
                        padded = [str(c or "").strip() for c in row]
                        if len(padded) < len(headers):
                            padded += [""] * (len(headers) - len(padded))
                        all_rows.append(padded[: len(headers)])
                if all_rows:
                    break   # encontrou tabela válida, não precisa tentar a próxima estratégia

    if all_rows and headers:
        df = pd.DataFrame(all_rows, columns=headers)
        df = df.dropna(how="all").reset_index(drop=True)
        df = df.loc[:, (df != "").any()]
        return df

    # ── Camada 2: texto linha a linha (pdfplumber) ────────────────────────────
    text_lines: list = []
    with pdfplumber.open(io.BytesIO(conteudo)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            text_lines.extend(text.splitlines())

    ano_ref = _inferir_ano(text_lines)
    rows = _parse_linhas_transacao(text_lines, ano_referencia=ano_ref)
    if rows:
        return pd.DataFrame(rows)

    # ── Camada 3: PyMuPDF (engine alternativa, decodifica fontes customizadas) ─
    mupdf_lines = _extrair_texto_pymupdf(conteudo)
    if mupdf_lines:
        ano_ref = _inferir_ano(mupdf_lines) or ano_ref
        rows = _parse_linhas_transacao(mupdf_lines, ano_referencia=ano_ref)
        if rows:
            return pd.DataFrame(rows)

    # ── Camada 4: OCR (Tesseract) ─────────────────────────────────────────────
    # Pode lançar ValueError se Tesseract não estiver instalado
    df_ocr = _parsear_pdf_ocr(conteudo)
    if not df_ocr.empty:
        return df_ocr

    raise ValueError(
        "Não foi possível extrair transações deste PDF. "
        "O arquivo parece ser um PDF baseado em imagem (escaneado). "
        "Se o OCR não está ativo no servidor, faça o redeploy no Render. "
        "Alternativa: exporte a fatura como CSV ou Excel diretamente no internet banking do banco."
    )


# ── DETECÇÃO DE FORMATO ──────────────────────────────────────────────────────

def _ler_dataframe(conteudo: bytes, nome: str) -> pd.DataFrame:
    """Lê CSV, Excel ou PDF e retorna DataFrame com colunas normalizadas (sem espaços extras)."""
    ext = nome.lower().split(".")[-1]

    if ext == "pdf":
        df = _parsear_pdf(conteudo)
        df.columns = [str(c).strip() for c in df.columns]
        return df

    buf = io.BytesIO(conteudo)
    if ext == "csv":
        for sep in [";", ",", "\t"]:
            try:
                buf.seek(0)
                df = pd.read_csv(buf, sep=sep, encoding="utf-8", dtype=str)
                if len(df.columns) > 1:
                    df.columns = [str(c).strip() for c in df.columns]
                    return df
            except Exception:
                continue
        buf.seek(0)
        df = pd.read_csv(buf, sep=";", encoding="latin-1", dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        return df
    else:
        buf.seek(0)
        try:
            df = pd.read_excel(buf, engine="openpyxl", dtype=str)
        except Exception:
            buf.seek(0)
            df = pd.read_excel(buf, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        return df


def _extrair_colunas_pdf_rapido(conteudo: bytes) -> list:
    """
    Detecção RÁPIDA de colunas em PDFs textuais (sem OCR, sem renderização).
    Só retorna colunas se a tabela parecer uma listagem de transações financeiras,
    ou seja, tiver ao menos uma coluna de VALOR e uma de DATA ou DESCRIÇÃO.
    Retorna [] para PDFs imagem ou tabelas de resumo (boleto, totais, etc.).
    """
    palavras_resumo = {
        'fatura anterior', 'total', 'saldo', 'pagamento',
        'limite', 'encargo', 'juros', 'vencimento',
    }
    # Keywords que indicam colunas de transação
    kw_data  = {'data', 'date', 'dt', 'competência', 'competencia',
                'lançamento', 'lancamento', 'movimento'}
    kw_desc  = {'descrição', 'descricao', 'description', 'histórico', 'historico',
                'estabelecimento', 'fornecedor', 'memo', 'transação', 'transacao',
                'beneficiario', 'portador'}
    kw_valor = {'valor', 'value', 'amount', 'r$', ' vl', 'débito', 'debito',
                'crédito', 'credito', 'compra'}

    def _tem(col: str, kws: set) -> bool:
        cl = col.lower()
        return any(kw in cl for kw in kws)

    try:
        with pdfplumber.open(io.BytesIO(conteudo)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables() or []
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    header = [str(c or '').strip() for c in table[0]]
                    header = [h for h in header if h]
                    if len(header) < 2:
                        continue
                    # Rejeita tabelas cujos headers são todos de resumo
                    if all(h.lower() in palavras_resumo or h == '' for h in header):
                        continue
                    # Só aceita se tiver coluna de valor E (data ou descrição)
                    tem_v = any(_tem(h, kw_valor) for h in header)
                    tem_d = any(_tem(h, kw_data)  for h in header)
                    tem_e = any(_tem(h, kw_desc)  for h in header)
                    if tem_v and (tem_d or tem_e):
                        return header
    except Exception:
        pass
    return []  # PDF imagem ou sem tabela de transações — OCR cuida


def detectar_colunas(conteudo: bytes, nome: str) -> list:
    """
    Retorna lista de colunas detectadas no arquivo.
    - CSV / Excel: retorna colunas do DataFrame.
    - PDF textual: retorna colunas da tabela via pdfplumber (rápido, sem OCR).
    - PDF imagem (scaneado/Santander): retorna [] — OCR cuida automaticamente.
    """
    ext = nome.lower().split(".")[-1]
    if ext == "pdf":
        return _extrair_colunas_pdf_rapido(conteudo)
    df = _ler_dataframe(conteudo, nome)
    return [str(c).strip() for c in df.columns if str(c).strip()]


# ── NORMALIZAÇÃO DE SAÍDA PDF ────────────────────────────────────────────────

def _normalizar_fatura_pdf(df: pd.DataFrame) -> pd.DataFrame:
    """
    Converte DataFrame bruto retornado pelo parser PDF para o formato padrão
    da conciliação: valor=float, data=datetime, cartao=str, tipo=str.

    As camadas 2/3/4 do _parsear_pdf() retornam strings ('15/03/2024', '1.500,00')
    que precisam ser convertidas antes de chegar ao conciliacao.py.
    Sem isso: pandas.sum() em strings faz concatenação → total_fatura = 0.
    """
    if df.empty:
        return df

    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    if 'valor' not in df.columns:
        return pd.DataFrame()

    result = pd.DataFrame()
    result['valor'] = _limpar_valor_col(df['valor']).abs().round(2)

    if 'data' in df.columns:
        result['data'] = pd.to_datetime(
            df['data'].astype(str), dayfirst=True, errors='coerce'
        )
    else:
        result['data'] = pd.NaT

    result['descricao'] = (
        df['descricao'].fillna('').astype(str)
        if 'descricao' in df.columns else ''
    )
    result['cartao'] = (
        df['cartao'].astype(str) if 'cartao' in df.columns else 'Fatura'
    )
    result['tipo'] = (
        df['tipo'].astype(str) if 'tipo' in df.columns else 'compra'
    )

    mask = result['valor'] > 0
    return result[mask].reset_index(drop=True)


# ── PARSER FATURA CARTÃO ─────────────────────────────────────────────────────

def parsear_fatura_cartao(conteudo: bytes, nome: str,
                          mapeamento: Optional[dict] = None) -> pd.DataFrame:
    """
    Parseia fatura de cartão de crédito.
    Suporta CSV, Excel (.xlsx/.xls), OFX/QFX e PDF.

    Estratégia:
      1. OFX/QFX  → parser SGML dedicado
      2. Com mapeamento explícito → usa colunas informadas pelo usuário
      3. XLSX/XLS → tenta formato Bradesco; se vazio, cai no parser genérico
      4. CSV/PDF  → parser genérico com detecção automática de colunas
    """
    ext = nome.lower().split(".")[-1]

    # ── OFX / QFX ─────────────────────────────────────────────────────────────
    if ext in ["ofx", "qfx"]:
        df = _parsear_ofx(conteudo)
        if not df.empty:
            df["cartao"] = "OFX"
            df["tipo"]   = "compra"
            return df[["data", "descricao", "valor", "cartao", "tipo"]]
        raise ValueError("Arquivo OFX não contém transações reconhecíveis.")

    # ── PDF: extrai via pdfplumber/OCR e normaliza para formato padrão ───────
    if ext == "pdf":
        df = _parsear_pdf(conteudo)
        if df.empty:
            raise ValueError(
                "Não foi possível extrair transações do PDF. "
                "Verifique se o arquivo está correto."
            )

        # Se o usuário mapeou colunas (PDF textual com tabela visível)
        if mapeamento:
            return _parsear_fatura_generico(df, mapeamento)

        # Camadas 2/3/4 produzem colunas padronizadas (data, descricao, valor)
        # mas ainda como strings — precisa normalizar tipos.
        colunas_std = {'data', 'descricao', 'valor'}
        if colunas_std.issubset(set(df.columns)):
            return _normalizar_fatura_pdf(df)

        # Camada 1 (tabela pdfplumber) → colunas arbitrárias → auto-detecção
        result = _parsear_fatura_csv(df)
        if not result.empty:
            return result

        # Fallback: normaliza o que tiver
        return _normalizar_fatura_pdf(df)

    # ── Mapeamento explícito (usuário mapeou as colunas) ──────────────────────
    if mapeamento:
        df = _ler_dataframe(conteudo, nome)
        return _parsear_fatura_generico(df, mapeamento)

    # ── XLSX/XLS: tenta Bradesco; se não encontrar, cai no genérico ───────────
    if ext in ["xlsx", "xls"]:
        try:
            wb = load_workbook(io.BytesIO(conteudo), read_only=True)
            ws = wb.active
            df_bd = _parsear_fatura_xlsx(ws)
            if not df_bd.empty:
                return df_bd
        except Exception:
            pass
        # Fallback genérico para qualquer Excel tabular
        df = _ler_dataframe(conteudo, nome)
        return _parsear_fatura_csv(df)

    # ── CSV ───────────────────────────────────────────────────────────────────
    df = _ler_dataframe(conteudo, nome)
    return _parsear_fatura_csv(df)


def _parsear_fatura_xlsx(ws) -> pd.DataFrame:
    skip_kw = ['total', 'transações', 'demonstrativo', 'data', 'cotação', 'pagamento de fatura']
    rows = []
    current_cartao = None

    for row in ws.iter_rows(values_only=True):
        col0 = str(row[0]).strip() if row[0] else ''
        # Tenta localizar valor na coluna R$ (índice 7) ou última coluna numérica
        valor = None
        for cell in reversed(row):
            try:
                v = float(cell)
                if v > 0:
                    valor = v
                    break
            except (TypeError, ValueError):
                continue

        # Detecta titular do cartão
        if re.search(r'\d{4} XXXX XXXX \d{4}', col0) or col0.startswith('@ '):
            m = re.search(r'(\d{4})$', col0.replace(' ', ''))
            current_cartao = m.group(1) if m else col0[-4:]
            continue

        # Encargos: juros, multa, IOF
        if col0.startswith('(+)') and valor:
            desc = col0.replace('(+)', '').strip()
            rows.append({'data': None, 'descricao': desc, 'valor': round(valor, 2),
                         'cartao': current_cartao, 'tipo': 'encargo'})
            continue

        if any(k in col0.lower() for k in skip_kw):
            continue
        if not valor:
            continue

        # Transação: data + descrição na mesma célula
        m = re.match(r'(\d{2}-\d{2}-\d{4})\s+(.+)', col0)
        if m:
            try:
                dt = pd.to_datetime(m.group(1), format='%d-%m-%Y')
            except Exception:
                dt = None
            rows.append({'data': dt, 'descricao': m.group(2).strip(),
                         'valor': round(valor, 2), 'cartao': current_cartao, 'tipo': 'compra'})

    return pd.DataFrame(rows)


def _parsear_fatura_csv(df: pd.DataFrame) -> pd.DataFrame:
    """
    Parseia fatura em formato tabular (CSV / Excel genérico / PDF).
    Detecta automaticamente colunas de data, descrição e valor.
    Cobre nomenclaturas de diversos bancos e ERPs.
    """
    col_map = _mapear_colunas_automatico(df, {
        'data': [
            'data', 'date', 'dt', 'data lançamento', 'data transação',
            'data da transação', 'data compra', 'data de compra',
            'data pagamento', 'data de pagamento', 'vencimento',
            'data vencimento', 'data do lançamento', 'data movimento',
        ],
        'descricao': [
            'descrição', 'descricao', 'description', 'historico', 'histórico',
            'estabelecimento', 'fornecedor', 'lançamento', 'lancamento',
            'transação', 'transacao', 'detalhe', 'memo', 'nome',
            'portador', 'comercio', 'comércio', 'local', 'beneficiario',
        ],
        'valor': [
            'valor', 'value', 'amount', 'r$', 'vl', 'vlr',
            'valor r$', 'valor (r$)', 'valor brl', 'importe',
            'valor nacional', 'valor original', 'valor da compra',
            'valor transação', 'valor transacao', 'débito', 'debito',
        ],
        'cartao': [
            'cartão', 'cartao', 'card', 'final', 'últimos dígitos',
            'portador', 'titular',
        ],
    })

    col_v = col_map.get('valor', '')
    if not col_v or col_v not in df.columns:
        return pd.DataFrame()

    valor = _limpar_valor_col(df[col_v])
    mask  = valor.abs() > 0
    df_f  = df[mask].copy()

    if df_f.empty:
        return pd.DataFrame()

    result            = pd.DataFrame()
    result['valor']   = valor[mask].abs().round(2).values
    result['tipo']    = 'compra'

    col_cartao = col_map.get('cartao', '')
    result['cartao'] = (
        df_f[col_cartao].astype(str).values
        if col_cartao and col_cartao in df_f.columns else 'Fatura'
    )

    col_d = col_map.get('data', '')
    result['data'] = (
        pd.to_datetime(df_f[col_d].astype(str), dayfirst=True, errors='coerce').values
        if col_d and col_d in df_f.columns else pd.NaT
    )

    col_desc = col_map.get('descricao', '')
    result['descricao'] = (
        df_f[col_desc].astype(str).values
        if col_desc and col_desc in df_f.columns else ''
    )
    return result


def _parsear_fatura_generico(df: pd.DataFrame, mapeamento: dict) -> pd.DataFrame:
    """
    Parseia fatura usando mapeamento explícito de colunas informado pelo usuário.
    Suporta qualquer banco/formato.
    """
    # Garante que as colunas do df estão sem espaços extras
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # Resolve o nome real da coluna no df (tolerante a case e espaço)
    col_lower = {c.lower(): c for c in df.columns}

    def _resolve(key: str) -> str:
        """Retorna o nome real da coluna ou '' se não encontrada."""
        v = str(mapeamento.get(key, '') or '').strip()
        if not v:
            return ''
        if v in df.columns:
            return v
        # tentativa case-insensitive
        return col_lower.get(v.lower(), '')

    col_v = _resolve('valor')
    if not col_v:
        return pd.DataFrame()

    valor = _limpar_valor_col(df[col_v])
    # Mantém valores positivos (compras) e também abs() de negativos — descarta só zeros
    mask  = valor.abs() > 0
    df_f  = df[mask].copy()

    if df_f.empty:
        return pd.DataFrame()

    result          = pd.DataFrame()
    result['valor'] = valor[mask].abs().round(2).values   # fatura: valor sempre positivo
    result['tipo']  = 'compra'

    col_cartao = _resolve('cartao')
    result['cartao'] = (
        df_f[col_cartao].astype(str).values
        if col_cartao else 'Fatura'
    )

    col_d = _resolve('data')
    result['data'] = (
        pd.to_datetime(df_f[col_d].astype(str), dayfirst=True, errors='coerce').values
        if col_d else pd.NaT
    )

    col_desc = _resolve('descricao')
    result['descricao'] = (
        df_f[col_desc].astype(str).values
        if col_desc else ''
    )
    return result


# ── UTILITÁRIO DE LIMPEZA DE VALOR ──────────────────────────────────────────

def _limpar_valor_col(series: pd.Series) -> pd.Series:
    """
    Converte coluna de texto para float.
    Suporta formato BR (1.500,00), internacional (1,500.00) e simples (1500.00 / 1500,00).
    Remove R$, espaços, parênteses contábeis.
    """
    def _parse(v) -> float:
        if pd.isna(v):
            return 0.0
        s = re.sub(r'[R$\s\xa0]', '', str(v)).strip()
        if not s or s in ('-', 'nan', 'None', ''):
            return 0.0

        # Valor negativo em notação contábil: (1.500,00) ou -1500,00
        neg = s.startswith('-') or (s.startswith('(') and s.endswith(')'))
        s = s.lstrip('(').rstrip(')').lstrip('-')

        has_dot   = '.' in s
        has_comma = ',' in s

        if has_dot and has_comma:
            # Ambos os separadores → o último é o decimal
            if s.rfind(',') > s.rfind('.'):
                # Formato BR: 1.500,00
                s = s.replace('.', '').replace(',', '.')
            else:
                # Formato INT: 1,500.00
                s = s.replace(',', '')
        elif has_comma:
            # Só vírgula → separador decimal (1500,00 ou 1,50)
            s = s.replace(',', '.')
        # else: só ponto ou nenhum → já está no formato correto

        try:
            result = float(s)
            return -result if neg else result
        except ValueError:
            return 0.0

    return series.map(_parse)


# ── PARSER ERP (genérico com mapeamento) ─────────────────────────────────────

DEFAULT_MAPEAMENTO_ERP = {
    'data':          'Data de Vencimento',
    'descricao':     'Descrição da Conta',
    'valor':         'Valor Liquidado',
    'valor_fallback':'Valor da Conta',
    'numero_fatura': 'Numero da Conta',
    'status':        'Status da Conta',
    'observacao':    'Observação da Conta',
}


def parsear_erp(conteudo: bytes, nome: str, mapeamento: Optional[dict] = None) -> pd.DataFrame:
    """
    Parseia exportação do ERP de contas a pagar/receber.
    Operações vetorizadas — sem iterrows, muito mais rápido para arquivos grandes.
    """
    df   = _ler_dataframe(conteudo, nome)
    mapa = {**DEFAULT_MAPEAMENTO_ERP, **(mapeamento or {})}

    # ── Valor: coluna principal, depois fallback ──────────────────────────────
    col_v  = mapa.get('valor', '')
    col_fb = mapa.get('valor_fallback', '')

    v1 = _limpar_valor_col(df[col_v])  if (col_v  and col_v  in df.columns) else pd.Series(0.0, index=df.index)
    v2 = _limpar_valor_col(df[col_fb]) if (col_fb and col_fb in df.columns) else pd.Series(0.0, index=df.index)

    # Usa v1 onde > 0, senão v2
    valor = v1.where(v1 > 0, v2)
    mask  = valor > 0
    df_f  = df[mask].copy()

    if df_f.empty:
        return pd.DataFrame(columns=['data', 'descricao', 'valor', 'numero_fatura', 'status', 'observacao'])

    result = pd.DataFrame(index=range(mask.sum()))
    result['valor'] = valor[mask].round(2).values

    # ── Data ─────────────────────────────────────────────────────────────────
    col_d = mapa.get('data', '')
    result['data'] = (
        pd.to_datetime(df_f[col_d].astype(str), dayfirst=True, errors='coerce').values
        if col_d and col_d in df_f.columns else pd.NaT
    )

    # ── Strings ──────────────────────────────────────────────────────────────
    for field in ['descricao', 'numero_fatura', 'status', 'observacao', 'categoria']:
        col = mapa.get(field, '')
        result[field] = (
            df_f[col].astype(str).str.strip().values
            if col and col in df_f.columns else ''
        )

    return result


# ── PARSERS RECEITAS ─────────────────────────────────────────────────────────

def parsear_operadora(conteudo: bytes, nome: str) -> pd.DataFrame:
    """Parseia extrato de operadora (Stone, Cielo, Rede…) — CSV, Excel ou OFX. Vetorizado."""
    ext = nome.lower().split(".")[-1]
    if ext in ["ofx", "qfx"]:
        df = _parsear_ofx(conteudo)
        if not df.empty:
            return df[["data", "descricao", "valor", "nsu", "tipo"]]

    df = _ler_dataframe(conteudo, nome)
    col_map = _mapear_colunas_automatico(df, {
        'data':    ['data', 'date', 'data venda', 'data transacao'],
        'descricao':['descrição', 'descricao', 'tipo', 'bandeira'],
        'valor':   ['valor bruto', 'valor', 'amount', 'vl bruto', 'bruto'],
        'nsu':     ['nsu', 'tid', 'autorização', 'autorizacao'],
    })

    col_v = col_map.get('valor', '')
    if not col_v or col_v not in df.columns:
        return pd.DataFrame()

    valor = _limpar_valor_col(df[col_v])
    mask  = valor > 0
    df_f  = df[mask].copy()

    result = pd.DataFrame()
    result['valor'] = valor[mask].round(2).values
    result['tipo']  = 'venda'

    col_d = col_map.get('data', '')
    result['data'] = (
        pd.to_datetime(df_f[col_d].astype(str), dayfirst=True, errors='coerce').values
        if col_d and col_d in df_f.columns else pd.NaT
    )

    col_desc = col_map.get('descricao', '')
    result['descricao'] = (
        df_f[col_desc].astype(str).values
        if col_desc and col_desc in df_f.columns else ''
    )

    col_nsu = col_map.get('nsu', '')
    result['nsu'] = (
        df_f[col_nsu].astype(str).values
        if col_nsu and col_nsu in df_f.columns else ''
    )
    return result


def parsear_banco(conteudo: bytes, nome: str, mapeamento: Optional[dict] = None) -> pd.DataFrame:
    """Parseia extrato bancário de créditos — CSV, Excel ou OFX. Vetorizado."""
    ext = nome.lower().split(".")[-1]
    if ext in ["ofx", "qfx"]:
        df = _parsear_ofx(conteudo)
        if not df.empty:
            return df[["data", "descricao", "valor"]].copy()

    df = _ler_dataframe(conteudo, nome)
    col_map = _mapear_colunas_automatico(df, {
        'data':    ['data', 'date', 'data lançamento', 'lançamento'],
        'descricao':['histórico', 'historico', 'descrição', 'descricao'],
        'valor':   ['crédito', 'credito', 'valor', 'entrada'],
    })
    col_map.update(mapeamento or {})

    col_v = col_map.get('valor', '')
    if not col_v or col_v not in df.columns:
        return pd.DataFrame()

    valor = _limpar_valor_col(df[col_v])
    mask  = valor > 0
    df_f  = df[mask].copy()

    result = pd.DataFrame()
    result['valor'] = valor[mask].round(2).values

    col_d = col_map.get('data', '')
    result['data'] = (
        pd.to_datetime(df_f[col_d].astype(str), dayfirst=True, errors='coerce').values
        if col_d and col_d in df_f.columns else pd.NaT
    )

    col_desc = col_map.get('descricao', '')
    result['descricao'] = (
        df_f[col_desc].astype(str).values
        if col_desc and col_desc in df_f.columns else ''
    )
    return result


# ── UTILIDADE ────────────────────────────────────────────────────────────────

def _mapear_colunas_automatico(df: pd.DataFrame, campos: dict) -> dict:
    """Tenta mapear automaticamente colunas do DataFrame para campos esperados."""
    colunas_lower = {c.lower().strip(): c for c in df.columns}
    resultado = {}
    for campo, candidatos in campos.items():
        for cand in candidatos:
            if cand.lower() in colunas_lower:
                resultado[campo] = colunas_lower[cand.lower()]
                break
    return resultado
